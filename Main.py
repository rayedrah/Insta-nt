import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import yt_dlp
import requests
import threading
import time
import os
import re
import json

BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
PROFILE_DIR = os.path.join(BASE_DIR, "chrome_profile")
OUTPUT_DIR  = os.path.join(BASE_DIR, "accounts")

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ── driver ────────────────────────────────────────────────────────────────────

def get_driver():
    options = uc.ChromeOptions()

    options.add_argument("--headless=new")  # modern Chrome headless
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")



    options.add_argument(f"--user-data-dir={PROFILE_DIR}")
    return uc.Chrome(version_main=146, options=options)


def is_logged_in(driver):
    driver.get("https://www.instagram.com/")
    time.sleep(3)
    return "Login" not in driver.title and "Log in" not in driver.page_source[:500]


# ── download functions ────────────────────────────────────────────────────────

def download_media_ytdlp(post_url, media_dir, filename_base):
    """
    Downloads media using yt-dlp (videos only - images fail with "No video formats").
    Returns (local_path, media_type) or (None, "image_needed") if it's an image post.
    """
    out_template = os.path.join(media_dir, f"{filename_base}.%(ext)s")

    ydl_opts = {
        "outtmpl": out_template,
        "quiet": True,
        "no_warnings": True,
        "format": "bestvideo+bestaudio/best",
        "merge_output_format": "mp4",
        "cookiesfrombrowser": ("chrome", PROFILE_DIR, None, None),
        "retries": 3,
        "fragment_retries": 3,
    }

    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(post_url, download=True)

        ext = info.get("ext", "mp4")
        local_path = os.path.join(media_dir, f"{filename_base}.{ext}")

        # Try to find the downloaded file
        if not os.path.exists(local_path):
            local_path = os.path.join(media_dir, f"{filename_base}.mp4")

        if not os.path.exists(local_path):
            for fname in os.listdir(media_dir):
                if fname.startswith(filename_base):
                    local_path = os.path.join(media_dir, fname)
                    ext = fname.rsplit(".", 1)[-1]
                    break

        if not os.path.exists(local_path):
            return None, None

        media_type = "video" if ext in ("mp4", "mkv", "webm", "mov") else "image"
        return local_path, media_type

    except yt_dlp.utils.DownloadError as e:
        error_msg = str(e)
        if "No video formats found" in error_msg or "No formats found" in error_msg:
            # This is an image post, signal to use Selenium download
            return None, "image_needed"
        return None, None
    except Exception:
        return None, None


def download_image_selenium(driver, post_url, media_dir, filename_base):
    """
    Downloads images using Selenium to extract URLs (authenticated session).
    Returns (local_path, "image") or (None, None) on failure.
    """
    try:
        driver.get(post_url)
        time.sleep(3)
        
        page_source = driver.page_source
        img_urls = []
        
        # Pattern 1: og:image meta tag
        try:
            meta_imgs = driver.find_elements(By.CSS_SELECTOR, 'meta[property="og:image"]')
            for meta in meta_imgs:
                url = meta.get_attribute("content")
                if url:
                    img_urls.append(url)
        except:
            pass
        
        # Pattern 2: JSON display_url
        matches = re.findall(r'"display_url":"(https://[^"]+)"', page_source)
        img_urls.extend(matches)
        
        # Pattern 3: img tags
        try:
            imgs = driver.find_elements(By.CSS_SELECTOR, 'article img[src*="instagram"]')
            for img in imgs:
                src = img.get_attribute("src")
                if src and "s640x640" not in src:
                    img_urls.append(src)
        except:
            pass
        
        if not img_urls:
            return None, None
        
        # Download the first/highest quality image
        img_url = img_urls[0].replace("\\/", "/").replace("\\u0026", "&")
        
        response = requests.get(img_url, timeout=30)
        response.raise_for_status()
        
        # Determine extension
        ext = "jpg"
        if ".png" in img_url or "image/png" in response.headers.get("Content-Type", ""):
            ext = "png"
        
        local_path = os.path.join(media_dir, f"{filename_base}.{ext}")
        
        with open(local_path, "wb") as f:
            f.write(response.content)
        
        return local_path, "image"
        
    except Exception:
        return None, None


# ── post URL collection ───────────────────────────────────────────────────────

def collect_post_urls(driver, username):
    """
    Scrolls profile and collects ALL post URLs with progress tracking.
    """
    driver.get(f"https://www.instagram.com/{username}/")
    time.sleep(5)

    if "Page Not Found" in driver.page_source or "Sorry, this page" in driver.page_source:
        print(f"[!] Profile '{username}' not found.")
        return []

    post_urls = set()
    last_count = 0
    stall_count = 0

    print(f"[*] Collecting post URLs for @{username} ...")

    while True:
        # Find all post/reel links
        anchors = driver.find_elements(By.CSS_SELECTOR, "a[href*='/p/'], a[href*='/reel/']")
        for a in anchors:
            href = a.get_attribute("href")
            if href:
                href = re.sub(r"\?.*", "", href.rstrip("/")) + "/"
                post_urls.add(href)

        current_count = len(post_urls)
        
        if current_count > last_count:
            print(f"    Found {current_count} posts so far...")
            last_count = current_count
            stall_count = 0
        else:
            stall_count += 1

        # Stop after 5 scrolls with no new posts
        if stall_count >= 5:
            break

        # Scroll aggressively
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2.0)
        
        # Extra small scrolls to trigger lazy loading
        for _ in range(2):
            driver.execute_script("window.scrollBy(0, 500);")
            time.sleep(0.3)

    print(f"[*] Found {len(post_urls)} unique post URLs.\n")
    return list(post_urls)


# ── metadata extraction ───────────────────────────────────────────────────────

def scrape_post(driver, url):
    """
    Extracts metadata from post using JSON-LD and multiple fallback methods.
    """
    driver.get(url)
    time.sleep(3)

    data = {
        "url": url,
        "username": None,
        "caption": None,
        "likes": None,
        "comments": None,
        "date": None,
        "media_type": None,
    }
    
    page_source = driver.page_source

    # ── Strategy 1: Extract from JSON-LD (most reliable) ──
    try:
        json_match = re.search(r'<script type="application/ld\+json">(.+?)</script>', page_source, re.DOTALL)
        if json_match:
            json_data = json.loads(json_match.group(1))
            
            # Username
            if "author" in json_data and isinstance(json_data["author"], dict):
                data["username"] = json_data["author"].get("alternateName") or json_data["author"].get("name")
            
            # Caption
            if "articleBody" in json_data:
                data["caption"] = json_data["articleBody"]
            elif "caption" in json_data:
                data["caption"] = json_data["caption"]
            
            # Engagement
            if "interactionStatistic" in json_data:
                for stat in json_data["interactionStatistic"]:
                    interaction_type = stat.get("interactionType", "")
                    if "LikeAction" in interaction_type:
                        data["likes"] = str(stat.get("userInteractionCount", ""))
                    elif "CommentAction" in interaction_type:
                        data["comments"] = str(stat.get("userInteractionCount", ""))
            
            # Date
            if "uploadDate" in json_data:
                data["date"] = json_data["uploadDate"]
    except:
        pass

    # ── Strategy 2: Extract from shared data JSON (Instagram's internal data) ──
    if not data["caption"] or not data["likes"]:
        try:
            # Look for window._sharedData or similar structures
            shared_data_match = re.search(r'window\._sharedData\s*=\s*({.+?});</script>', page_source, re.DOTALL)
            if shared_data_match:
                shared_json = json.loads(shared_data_match.group(1))
                
                # Navigate through the nested structure to find post data
                entry_data = shared_json.get("entry_data", {})
                post_page = entry_data.get("PostPage", [{}])[0]
                media = post_page.get("graphql", {}).get("shortcode_media", {})
                
                if not data["caption"]:
                    edges = media.get("edge_media_to_caption", {}).get("edges", [])
                    if edges:
                        data["caption"] = edges[0].get("node", {}).get("text", "")
                
                if not data["likes"]:
                    data["likes"] = str(media.get("edge_media_preview_like", {}).get("count", ""))
                
                if not data["comments"]:
                    data["comments"] = str(media.get("edge_media_to_comment", {}).get("count", ""))
        except:
            pass

    # ── Strategy 3: Regex patterns on page source ──
    if not data["username"]:
        try:
            url_match = re.search(r'instagram\.com/([^/]+)/', url)
            if url_match:
                data["username"] = url_match.group(1)
        except:
            pass

    if not data["caption"]:
        try:
            # Try to find caption in various patterns
            caption_patterns = [
                r'"edge_media_to_caption":\s*{\s*"edges":\s*\[\s*{\s*"node":\s*{\s*"text":\s*"([^"]+)"',
                r'<meta property="og:description" content="([^"]+)"',
                r'"caption":\s*"([^"]+)"',
            ]
            for pattern in caption_patterns:
                match = re.search(pattern, page_source, re.DOTALL)
                if match:
                    data["caption"] = match.group(1).replace("\\n", "\n").replace('\\"', '"')
                    break
        except:
            pass

    # Try to extract caption from visible h1 elements as last resort
    if not data["caption"]:
        try:
            h1_elements = driver.find_elements(By.CSS_SELECTOR, "h1")
            for h1 in h1_elements:
                text = h1.text.strip()
                if text and len(text) > 5 and text != data["username"]:
                    data["caption"] = text
                    break
        except:
            pass

    if not data["likes"]:
        try:
            patterns = [
                r'"edge_media_preview_like":\s*{\s*"count":\s*(\d+)',
                r'"edge_liked_by":\s*{\s*"count":\s*(\d+)',
                r'(\d+)\s+likes?',
            ]
            for pattern in patterns:
                m = re.search(pattern, page_source)
                if m:
                    data["likes"] = m.group(1)
                    break
        except:
            pass

    if not data["comments"]:
        try:
            patterns = [
                r'"edge_media_to_comment":\s*{\s*"count":\s*(\d+)',
                r'"edge_media_preview_comment":\s*{\s*"count":\s*(\d+)',
                r'(\d+)\s+comments?',
            ]
            for pattern in patterns:
                m = re.search(pattern, page_source)
                if m:
                    data["comments"] = m.group(1)
                    break
        except:
            pass

    if not data["date"]:
        try:
            time_elem = driver.find_element(By.CSS_SELECTOR, "time[datetime]")
            data["date"] = time_elem.get_attribute("datetime")
        except:
            try:
                date_match = re.search(r'"uploadDate":"([^"]+)"', page_source)
                if date_match:
                    data["date"] = date_match.group(1)
            except:
                pass

    # Media type detection
    try:
        driver.find_element(By.CSS_SELECTOR, "video")
        data["media_type"] = "video"
    except:
        data["media_type"] = "image"

    return data


# ── Excel helpers ─────────────────────────────────────────────────────────────

HEADERS    = ["URL", "Username", "Caption", "Likes", "Comments",
              "Date", "Media File", "Media Type", "Scraped At"]
COL_WIDTHS = [50, 20, 60, 10, 12, 25, 45, 12, 22]

HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
ROW_FONT    = Font(name="Arial", size=10)
LINK_FONT   = Font(name="Arial", size=10, color="0563C1", underline="single")
ALT_FILL    = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")


def excel_path(username):
    return os.path.join(OUTPUT_DIR, f"{username}.xlsx")


def get_media_dir(username):
    path = os.path.join(OUTPUT_DIR, f"{username}-media")
    os.makedirs(path, exist_ok=True)
    return path


def setup_excel(username):
    path = excel_path(username)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = username[:31]

        for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A2"
        wb.save(path)
        print(f"[*] Created {path}\n")
    else:
        print(f"[*] Appending to {path}\n")
    return path


def save_to_excel(path, data, local_media_path):
    wb = load_workbook(path)
    ws = wb.active

    row_idx = ws.max_row + 1
    fill = ALT_FILL if row_idx % 2 == 0 else None
    scraped_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    row_values = [
        data["url"], data["username"], data["caption"],
        data["likes"], data["comments"], data["date"],
        None,
        data["media_type"], scraped_at,
    ]

    for col, val in enumerate(row_values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = ROW_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=(col == 3))
        if fill:
            cell.fill = fill

    media_cell = ws.cell(row=row_idx, column=7)
    if local_media_path:
        media_cell.value = os.path.basename(local_media_path)
        media_cell.hyperlink = local_media_path
        media_cell.font = LINK_FONT
    else:
        media_cell.value = "download failed"
        media_cell.font = ROW_FONT
    if fill:
        media_cell.fill = fill

    wb.save(path)


# ── Main loop ─────────────────────────────────────────────────────────────────

driver = get_driver()

if not is_logged_in(driver):
    print("[!] Not logged in — run Setup.py first.")
    driver.quit()
    exit()

print("[*] Logged in, using saved session.\n")

while True:
    username = input("Enter Instagram username (or 'exit'): ").strip().lstrip("@")

    if username.lower() == "exit":
        print("[*] Exiting.")
        break

    if not username:
        continue

    # Step 1: Collect all post URLs by scrolling
    urls = collect_post_urls(driver, username)

    if not urls:
        print(f"[!] No posts found for @{username}, skipping.\n")
        continue

    xl_path = setup_excel(username)
    mdir    = get_media_dir(username)

    print(f"[*] Starting parallel download of {len(urls)} posts...\n")
    
    # Step 2: Download all media in parallel (4 workers using tabs in single driver)
    download_results = {}  # url -> (local_path, media_type, shortcode)
    
    # Create a lock for thread-safe tab switching
    driver_lock = threading.Lock()
    
    # Open 4 tabs in the main driver for parallel downloads
    original_window = driver.current_window_handle
    tabs = [original_window]
    
    for _ in range(3):  # Create 3 more tabs (total 4)
        driver.switch_to.new_window('tab')
        tabs.append(driver.current_window_handle)
    
    # Switch back to original tab
    driver.switch_to.window(original_window)
    
    def download_with_tab(url, mdir, index, total, driver_ref, tab_handles, lock):
        """Downloads using a specific tab from the driver"""
        shortcode_match = re.search(r'/(p|reel)/([^/]+)/', url)
        shortcode = shortcode_match.group(2) if shortcode_match else f"post_{index}"
        
        print(f"  [{index}/{total}] Downloading: {url}")
        
        # Try yt-dlp first (no driver needed, no lock needed)
        local_path, media_type = download_media_ytdlp(url, mdir, shortcode)
        
        # If it's an image, use Selenium with assigned tab (thread-safe with lock)
        if media_type == "image_needed":
            # Use lock to prevent concurrent tab switching
            with lock:
                tab_idx = index % len(tab_handles)
                driver_ref.switch_to.window(tab_handles[tab_idx])
                local_path, media_type = download_image_selenium(driver_ref, url, mdir, shortcode)
        
        if local_path:
            print(f"  [{index}/{total}] ✓ Downloaded: {os.path.basename(local_path)}")
        else:
            print(f"  [{index}/{total}] ✗ Failed")
        
        return url, local_path, media_type, shortcode
    
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {
            executor.submit(download_with_tab, url, mdir, i+1, len(urls), driver, tabs, driver_lock): url
            for i, url in enumerate(urls)
        }
        
        for future in as_completed(futures):
            url, local_path, media_type, shortcode = future.result()
            download_results[url] = (local_path, media_type, shortcode)
    
    # Close extra tabs, keep original
    for tab in tabs[1:]:
        driver.switch_to.window(tab)
        driver.close()
    driver.switch_to.window(original_window)
    
    print(f"\n[*] Downloads complete. Extracting metadata...\n")
    
    # Step 3: Extract metadata serially (fast, just scraping HTML)
    for i, url in enumerate(urls, 1):
        print(f"  [{i}/{len(urls)}] Extracting metadata: {url}")
        
        try:
            post_data = scrape_post(driver, url)
            
            # Get download results
            local_path, detected_type, shortcode = download_results.get(url, (None, None, None))
            
            if detected_type and detected_type != "image_needed":
                post_data["media_type"] = detected_type
            
            save_to_excel(xl_path, post_data, local_path)
            
        except Exception as e:
            print(f"     [error] {str(e)[:50]}")
        
        time.sleep(1.5)  # Rate limiting

    print(f"\n[*] Done!")
    print(f"    Excel → {xl_path}")
    print(f"    Media → {mdir}\n")

driver.quit()
